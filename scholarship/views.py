import os, pytz

from django.contrib.auth import get_user_model
from django.core.files.storage import get_storage_class
from django.core.mail import EmailMessage
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse
from django.utils import timezone
from django.views import View
from django.views.generic import ListView
from django.views.generic.edit import CreateView

from braces.views import LoginRequiredMixin
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from .mixins import ModeratorsMixin

from formtools.wizard.views import SessionWizardView
from scholarship.forms import (
    PersonalInformationForm,
    HighSchoolForm,
    AcademicCounselorForm,
    CurrentEmploymentForm,
    ParentForm,
    HouseholdForm,
    CollegeForm,
    OtherForm,
    AttachmentForm,
    UserFilterForm,
)
from scholarship.models import ApplicationState, Attachments, TemporaryStorage


class ScholarshipView(LoginRequiredMixin, SessionWizardView):
    form_list = [
        PersonalInformationForm,
        HighSchoolForm,
        AcademicCounselorForm,
        CurrentEmploymentForm,
        ParentForm,
        HouseholdForm,
        CollegeForm,
        OtherForm,
        AttachmentForm,
    ]
    template_name = "scholarship/index.html"

    def get(self, request, *args, **kwargs):
        application_state = ApplicationState.objects.first()
        michigan_tz = pytz.timezone('America/Detroit')
        michigan_time = timezone.now().astimezone(michigan_tz)
        
        if application_state.start_date <= michigan_time.date() <= application_state.end_date:
            if self.request.user.has_submitted_application:
                return redirect("scholarship-application-attachments")
            
            return super().get(request, *args, **kwargs)
        else:
            return redirect("scholarship-application-closed")

    def get_form(self, step=None, data=None, files=None):
        if step is None:
            step = self.steps.current

        # Retrieve any existing data for this step
        temp_data = {}
        try:
            temp_storage = TemporaryStorage.objects.get(
                step=step, user=self.request.user
            )
            temp_data = temp_storage.data
            print(temp_data, data)
            # Merge the POST data and the TemporaryStorage data
            if data is not None:
                data = data
            else:
                data = temp_data
        except TemporaryStorage.DoesNotExist:
            pass

        return super().get_form(step, data, files)

    def process_step(self, form):
        if form.is_valid():
            data = form.data

        if data is None:
            data = {}

        TemporaryStorage.objects.update_or_create(
            step=self.steps.current, user=self.request.user, defaults={"data": data}
        )

        return super().process_step(form)

    def done(self, form_list, **kwargs):
        personal_information_form = form_list[0]
        if personal_information_form.cleaned_data:
            personal_information = personal_information_form.save(commit=False)
            personal_information.user = self.request.user
            personal_information.save()

        high_school_form = form_list[1]
        if high_school_form.cleaned_data:
            high_school = high_school_form.save(commit=False)
            high_school.user = self.request.user
            high_school.save()

        academic_counselor_form = form_list[2]
        if academic_counselor_form.cleaned_data:
            academic_counselor = academic_counselor_form.save(commit=False)
            academic_counselor.user = self.request.user
            academic_counselor.save()

        current_employment_form = form_list[3]
        if current_employment_form.cleaned_data:
            current_employment = current_employment_form.save(commit=False)
            current_employment.user = self.request.user
            current_employment.save()

        parent_form = form_list[4]
        if parent_form.cleaned_data:
            parent = parent_form.save(commit=False)
            parent.user = self.request.user
            parent.save()

        household_form = form_list[5]
        if household_form.cleaned_data:
            household = household_form.save(commit=False)
            household.user = self.request.user
            household.save()

        college_form = form_list[6]
        if college_form.cleaned_data:
            college = college_form.save(commit=False)
            college.user = self.request.user
            college.save()

        other_form = form_list[7]
        if other_form.cleaned_data:
            other = other_form.save(commit=False)
            other.user = self.request.user
            other.save()

        # Set the has_submitted_form flag in the user model to True
        self.request.user.has_submitted_application = True
        self.request.user.save()

        # Delete the TemporaryStorage records for the current user
        TemporaryStorage.objects.filter(user=self.request.user).delete()

        # Redirect the user to attachments page
        return redirect("scholarship-application-attachments")


class ScholarshipAttachmentView(LoginRequiredMixin, CreateView):
    template_name = "scholarship/attachments.html"
    model = Attachments
    form_class = AttachmentForm
    success_url = "/scholarship/application/success/"

    def get(self, request, *args, **kwargs):
        if not self.request.user.has_submitted_application:
            return redirect("scholarship-application")
        elif not self.request.user.has_submitted_attachments:
            return super().get(request, *args, **kwargs)
        else:
            return redirect("scholarship-application-success")

    def form_valid(self, form):
        form.instance.user = self.request.user
        self.request.user.has_submitted_attachments = True
        self.request.user.save()

        default_from_email = os.getenv("DEFAULT_FROM_EMAIL", default="")

        # Send confirmation email
        message = EmailMessage(
            subject="Thank You for Your Application to the Rosa Parks Scholarship Foundation",
            from_email=f"Rosa Parks Scholarship Foundation <{default_from_email}>",
            to=[self.request.user.email],
        )
        message.template_id = "scholarship form confirmation"
        message.merge_global_data = {
            "first_name": self.request.user.first_name,
            "last_name": self.request.user.last_name,
        }
        message.send()

        return super().form_valid(form)


class ScholarshipListView(LoginRequiredMixin, ModeratorsMixin, ListView):
    model = get_user_model()
    template_name = "scholarship/list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter_form"] = UserFilterForm(self.request.GET)
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        form = UserFilterForm(self.request.GET)

        if form.is_valid():
            if form.cleaned_data["completed_application"] == "true":
                queryset = queryset.filter(
                    has_submitted_application=True, has_submitted_attachments=True
                )
            elif form.cleaned_data["completed_application"] == "false":
                queryset = queryset.filter(
                    has_submitted_application=False, has_submitted_attachments=False
                )
            if form.cleaned_data["has_submitted_application"] == "true":
                queryset = queryset.filter(has_submitted_application=True)
            elif form.cleaned_data["has_submitted_application"] == "false":
                queryset = queryset.filter(has_submitted_application=False)
            if form.cleaned_data["has_submitted_attachments"] == "true":
                queryset = queryset.filter(has_submitted_attachments=True)
            elif form.cleaned_data["has_submitted_attachments"] == "false":
                queryset = queryset.filter(has_submitted_attachments=False)

        # Get the search query
        search_query = self.request.GET.get('search', '')

        # Filter the queryset based on the search query
        queryset = queryset.filter(
            Q(email__icontains=search_query)
        )

        return queryset


class ScholarshipDownloadExcelView(LoginRequiredMixin, ModeratorsMixin, View):
    def post(self, request, *args, **kwargs):
        user_ids = request.POST.getlist("user_ids")
        users = get_user_model().objects.filter(id__in=user_ids)
        storage = get_storage_class("page.storage_backends.PrivateMediaStorage")()

        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "First Name",
                "Last Name",
                "Email",
                "Address 1",
                "Address 2",
                "City",
                "State",
                "Zip Code",
                "Phone Number",
                "Date of Birth",
                "Place of Birth",
                "High School Name",
                "High School City",
                "High School Graduation Date",
                "High School GPA",
                "Academic Counselor Name",
                "Academic Counselor Phone Number",
                "Academic Counselor Email",
                "Current Employer Name",
                "Current Job Title",
                "Hours Per Week",
                "Parent 1 Full Name",
                "Parent 1 Email",
                "Parent 1 Address 1",
                "Parent 1 Address 2",
                "Parent 1 City",
                "Parent 1 State",
                "Parent 1 Zip Code",
                "Parent 1 Phone Number",
                "Parent 1 Place of Employment",
                "Parent 1 Job Title",
                "Parent 2 Full Name",
                "Parent 2 Email",
                "Parent 2 Address 1",
                "Parent 2 Address 2",
                "Parent 2 City",
                "Parent 2 State",
                "Parent 2 Zip Code",
                "Parent 2 Phone Number",
                "Parent 2 Place of Employment",
                "Parent 2 Job Title",
                "Total Household Income",
                "Number of Siblings Under 18",
                "Number of Siblings Over 18",
                "College Goal",
                "College Major",
                "College Career",
                "College Applied for #1",
                "College Applied for #2",
                "College Applied for #3",
                "College Plan to Attend",
                "College Savings",
                "College Savings by Guardian",
                "College Financial Need",
                "Foster Care",
                "Challenges",
                "Other Scholarships Applied For",
                "Other Scholarships Awarded",
                "Plan to Pay",
                "Reference Letter 1",
                "Reference Letter 2",
                "High School Transcript",
                "Honors and Awards",
                "Extracurricular Activities",
                "Community Service and Volunteer Activities",
                "Essay",
            ]
        )
        for user in users:
            ws.append(
                [
                    user.first_name,
                    user.last_name,
                    user.email,
                    getattr(user.personalinformation, "address1", "")
                    if hasattr(user, "personalinformation")
                    else "",
                    getattr(user.personalinformation, "address2", "")
                    if hasattr(user, "personalinformation")
                    else "",
                    getattr(user.personalinformation, "city", "")
                    if hasattr(user, "personalinformation")
                    else "",
                    getattr(user.personalinformation, "state", "")
                    if hasattr(user, "personalinformation")
                    else "",
                    getattr(user.personalinformation, "zip_code", "")
                    if hasattr(user, "personalinformation")
                    else "",
                    getattr(user.personalinformation, "phone_number.raw_input", "")
                    if hasattr(user, "personalinformation")
                    else "",
                    getattr(user.personalinformation, "dob", "")
                    if hasattr(user, "personalinformation")
                    else "",
                    getattr(user.personalinformation, "place_of_birth", "")
                    if hasattr(user, "personalinformation")
                    else "",
                    getattr(user.highschool, "name", "")
                    if hasattr(user, "highschool")
                    else "",
                    getattr(user.highschool, "city", "")
                    if hasattr(user, "highschool")
                    else "",
                    getattr(user.highschool, "graduation_date", "")
                    if hasattr(user, "highschool")
                    else "",
                    getattr(user.highschool, "gpa", "")
                    if hasattr(user, "highschool")
                    else "",
                    getattr(user.academiccounselor, "name", "")
                    if hasattr(user, "academiccounselor")
                    else "",
                    getattr(user.academiccounselor, "phone_number.raw_input", "")
                    if hasattr(user, "academiccounselor")
                    else "",
                    getattr(user.academiccounselor, "email", "")
                    if hasattr(user, "academiccounselor")
                    else "",
                    getattr(user.currentemployment, "name", "")
                    if hasattr(user, "currentemployment")
                    else "",
                    getattr(user.currentemployment, "job_title", "")
                    if hasattr(user, "currentemployment")
                    else "",
                    getattr(user.currentemployment, "hours_per_week", "")
                    if hasattr(user, "currentemployment")
                    else "",
                    getattr(user.parent, "parent_1_full_name", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_1_email", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_1_address1", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_1_address2", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_1_city", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_1_state", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_1_zip_code", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_1_phone_number.raw_input", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_1_place_of_employment", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_1_job_title", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_2_full_name", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_2_email", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_2_address1", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_2_address2", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_2_city", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_2_state", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_2_zip_code", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_2_phone_number.raw_input", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_2_place_of_employment", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.parent, "parent_2_job_title", "")
                    if hasattr(user, "parent")
                    else "",
                    getattr(user.household, "total_household_income", "")
                    if hasattr(user, "household")
                    else "",
                    getattr(user.household, "siblings_under_18", "")
                    if hasattr(user, "household")
                    else "",
                    getattr(user.household, "siblings_over_18", "")
                    if hasattr(user, "household")
                    else "",
                    getattr(user.college, "goal", "")
                    if hasattr(user, "college")
                    else "",
                    getattr(user.college, "major", "")
                    if hasattr(user, "college")
                    else "",
                    getattr(user.college, "career", "")
                    if hasattr(user, "college")
                    else "",
                    getattr(user.college, "applied_for_1", "")
                    if hasattr(user, "college")
                    else "",
                    getattr(user.college, "applied_for_2", "")
                    if hasattr(user, "college")
                    else "",
                    getattr(user.college, "applied_for_3", "")
                    if hasattr(user, "college")
                    else "",
                    getattr(user.college, "plan_to_attend", "")
                    if hasattr(user, "college")
                    else "",
                    getattr(user.college, "savings", "")
                    if hasattr(user, "college")
                    else "",
                    getattr(user.college, "savings_by_guardian", "")
                    if hasattr(user, "college")
                    else "",
                    getattr(user.college, "financial_need", "")
                    if hasattr(user, "college")
                    else "",
                    getattr(user.other, "foster_care", "")
                    if hasattr(user, "other")
                    else "",
                    getattr(user.other, "challenges", "")
                    if hasattr(user, "other")
                    else "",
                    getattr(user.other, "other_scholarships", "")
                    if hasattr(user, "other")
                    else "",
                    getattr(user.other, "other_scholarships_awarded", "")
                    if hasattr(user, "other")
                    else "",
                    getattr(user.other, "plan_to_pay", "")
                    if hasattr(user, "other")
                    else "",
                    storage.url(user.attachments.reference_letter_1.file.name)
                    if hasattr(user, "attachments") and user.attachments.reference_letter_1
                    else "",
                    storage.url(user.attachments.reference_letter_2.file.name)
                    if hasattr(user, "attachments") and user.attachments.reference_letter_2
                    else "",
                    storage.url(user.attachments.high_school_transcript.file.name)
                    if hasattr(user, "attachments") and user.attachments.high_school_transcript
                    else "",
                    storage.url(user.attachments.honors_awards.file.name)
                    if hasattr(user, "attachments") and user.attachments.honors_awards
                    else "",
                    storage.url(user.attachments.extracurricular_activities.file.name)
                    if hasattr(user, "attachments") and user.attachments.extracurricular_activities
                    else "",
                    storage.url(user.attachments.community_service_volunteer_activities.file.name)
                    if hasattr(user, "attachments") and user.attachments.community_service_volunteer_activities
                    else "",
                    storage.url(user.attachments.essay.file.name)
                    if hasattr(user, "attachments") and user.attachments.essay
                    else "",
                ]
            )
            
            # Get the cell with the URL and set it as a hyperlink
            url_reference_letter_1_cell = ws["{}{}".format(get_column_letter(60), ws.max_row)]
            url_reference_letter_2_cell = ws["{}{}".format(get_column_letter(61), ws.max_row)]
            url_high_school_transcript_cell = ws["{}{}".format(get_column_letter(62), ws.max_row)]  
            url_honors_awards_cell = ws["{}{}".format(get_column_letter(63), ws.max_row)]
            url_extracurricular_activities_cell = ws["{}{}".format(get_column_letter(64), ws.max_row)]
            url_community_service_volunteer_activities_cell = ws["{}{}".format(get_column_letter(65), ws.max_row)]
            url_essay_cell = ws["{}{}".format(get_column_letter(66), ws.max_row)]
            
            if hasattr(user, "attachments") and user.attachments.reference_letter_1:
                url_reference_letter_1_cell.hyperlink = storage.url(user.attachments.reference_letter_1.file.name)
                url_reference_letter_1_cell.style = "Hyperlink"
            
            if hasattr(user, "attachments") and user.attachments.reference_letter_2:
                url_reference_letter_2_cell.hyperlink = storage.url(user.attachments.reference_letter_2.file.name)
                url_reference_letter_2_cell.style = "Hyperlink"
            
            if hasattr(user, "attachments") and user.attachments.high_school_transcript:
                url_high_school_transcript_cell.hyperlink = storage.url(user.attachments.high_school_transcript.file.name)
                url_high_school_transcript_cell.style = "Hyperlink"
            
            if hasattr(user, "attachments") and user.attachments.honors_awards:
                url_honors_awards_cell.hyperlink = storage.url(user.attachments.honors_awards.file.name)
                url_honors_awards_cell.style = "Hyperlink"
            
            if hasattr(user, "attachments") and user.attachments.extracurricular_activities:
                url_extracurricular_activities_cell.hyperlink = storage.url(user.attachments.extracurricular_activities.file.name)
                url_extracurricular_activities_cell.style = "Hyperlink"
            
            if hasattr(user, "attachments") and user.attachments.community_service_volunteer_activities:
                url_community_service_volunteer_activities_cell.hyperlink = storage.url(user.attachments.community_service_volunteer_activities.file.name)
                url_community_service_volunteer_activities_cell.style = "Hyperlink"
            
            if hasattr(user, "attachments") and user.attachments.essay:
                url_essay_cell.hyperlink = storage.url(user.attachments.essay.file.name)
                url_essay_cell.style = "Hyperlink"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=scholarship-submissions.xlsx"
        wb.save(response)

        return response


class ScholarshipDeleteView(LoginRequiredMixin, ModeratorsMixin, ListView):
    model = get_user_model()
    template_name = "scholarship/delete.html"
    paginate_by = 25

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(
            has_submitted_application=True, has_submitted_attachments=True
        )

        # Get the search query
        search_query = self.request.GET.get('search', '')

        # Filter the queryset based on the search query
        queryset = queryset.filter(
            Q(email__icontains=search_query)
        )

        return queryset

    def post(self, request, *args, **kwargs):
        # Get the IDs of the selected submissions
        user_ids = request.POST.getlist('user_ids')
        users = get_user_model().objects.filter(id__in=user_ids)
        
        for user in users:            
            # Delete the user's application
            if hasattr(user, "personalinformation"):
                user.personalinformation.delete()
            if hasattr(user, "highschool"):
                user.highschool.delete()
            if hasattr(user, "academiccounselor"):
                user.academiccounselor.delete()
            if hasattr(user, "currentemployment"):
                user.currentemployment.delete()
            if hasattr(user, "parent"):
                user.parent.delete()
            if hasattr(user, "household"):
                user.household.delete()
            if hasattr(user, "college"):
                user.college.delete()
            if hasattr(user, "other"):
                user.other.delete()

            # Delete the user's attachments
            if hasattr(user, "attachments"):
                user.attachments.delete()
            
            user.has_submitted_application = False
            user.has_submitted_attachments = False
            user.save()

        # Return to scholarship delete page
        return redirect(reverse("scholarship-application-delete"))
